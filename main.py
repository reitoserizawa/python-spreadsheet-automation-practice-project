import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    product_num = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory_price = product_list.cell(product_row, 5)
    
    # calculate total products of each supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1
        
    # calculate total value of inventory per supplier
    if supplier_name in total_value_per_supplier and total_value_per_supplier:
        total_value_per_supplier[supplier_name] += (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price
        
    # return product number with inventory count that is under 10 inventory
    if inventory < 10:
        product_under_10_inv[int(product_num)] = int(inventory)
        
    # calculate total value of inventory per product number
    inventory_price.value = inventory * price
    
print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)
    
# create a new spreadsheet with updated columns
inv_file.save("inventory_with_total_value.xlsx")